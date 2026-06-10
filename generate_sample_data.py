"""Generate a synthetic property spreadsheet for testing the asset-group pipeline.

Schema mirrors the reference workbook, including the deliberate duplicate
'Property Type' header (shape vs form). Rows are entirely fabricated and
include edge cases the v1 skill cares about: blanks, integer-valued floats,
stray whitespace, mixed authorities, and varied age bands.
"""

import random
from openpyxl import Workbook

random.seed(42)

HEADERS = [
    "Property Reference", "Block Reference", "Complete Address", "Postcode",
    "Local Authority", "Property Type", "Property Type", "No. Bedrooms",
    "Construction Type", "Age Band", "Block Name", "Tenure", "Ward", "Region",
]

LOCAL_AUTHORITIES = [
    ("Coventry",   "West Midlands", ["CV1", "CV2", "CV3", "CV5"]),
    ("Birmingham", "West Midlands", ["B1",  "B12", "B15", "B29"]),
    ("Manchester", "North West",    ["M1",  "M14", "M20", "M22"]),
    ("Leeds",      "Yorkshire",     ["LS1", "LS6", "LS11","LS17"]),
    ("Accrington", "North West",    ["BB5"]),
    ("Bristol",    "South West",    ["BS1", "BS3", "BS5", "BS16"]),
]

SHAPE_TYPES = ["House", "Bungalow", "Flat", "Maisonette"]
FORM_TYPES  = ["H/B", "L/R F", "M/R F", "H/R F", "T/H"]
AGE_BANDS   = ["pre 1900", "1900 - 1918", "1919 - 1944",
               "1945 - 1964", "1965 - 1982", "1983 - 1999", "2000 - 2010", "post 2010"]
CONSTRUCTION = ["Trad", "Non-Trad", "System Build", "Steel Frame", "Timber Frame"]
TENURES = ["GN", "SH", "LH", "Shared Ownership"]
WARDS = ["Central", "North", "South", "East", "West", "Riverside", "Hillfield", None]
STREETS = ["FARMCOTE ROAD", "OAK AVENUE", "ELM CLOSE", "HIGH STREET",
           "STATION ROAD", "CHURCH LANE", "PARK VIEW", "MILL CRESCENT"]


def make_row(idx: int) -> list:
    la_name, region, postcode_areas = random.choice(LOCAL_AUTHORITIES)
    street = random.choice(STREETS)
    house_no = random.randint(1, 250)
    postcode = f"{random.choice(postcode_areas)} {random.randint(1,9)}{random.choice('ABDEFHJLNPQRSTUWXYZ')}{random.choice('ABDEFHJLNPQRSTUWXYZ')}"
    block_ref = random.choice([0.0, float(random.randint(100, 9999))])
    block_name = street if block_ref else None

    # Edge cases sprinkled in:
    address = f"{house_no}  {street}  "          # stray double-spaces & trailing spaces
    bedrooms = random.choice([1.0, 2.0, 3.0, 4.0, 5.0])  # integer-valued floats
    age_band = random.choice(AGE_BANDS)
    construction = random.choice(CONSTRUCTION)
    ward = random.choice(WARDS)                   # sometimes None (blank)

    # Roughly 1 in 25 rows: blank out an attribute to test the v1 blank-handling rule.
    if idx % 25 == 0:
        age_band = None
    if idx % 31 == 0:
        construction = None

    return [
        str(180000 + idx),          # Property Reference
        block_ref,                  # Block Reference
        address,                    # Complete Address (with stray whitespace)
        postcode,                   # Postcode
        la_name,                    # Local Authority
        random.choice(SHAPE_TYPES), # Property Type (shape)
        random.choice(FORM_TYPES),  # Property Type (form) — duplicate header
        bedrooms,                   # No. Bedrooms (float, e.g. 2.0)
        construction,               # Construction Type (occasionally blank)
        age_band,                   # Age Band (occasionally blank)
        block_name,                 # Block Name
        random.choice(TENURES),     # Tenure
        ward,                       # Ward (often blank)
        region,                     # Region
    ]


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(HEADERS)
    for i in range(1, 121):  # 120 synthetic rows — small enough to eyeball, big enough to group
        ws.append(make_row(i))
    out = "Sample property data.xlsx"
    wb.save(out)
    print(f"Wrote {out} with {ws.max_row - 1} rows and {ws.max_column} columns.")


if __name__ == "__main__":
    main()
